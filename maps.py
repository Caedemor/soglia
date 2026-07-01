"""
Hand-written ColumnMaps for the three real lists — stand-ins for what the LLM
(stage 1) will eventually emit. The generalized transcriber (stage 2) consumes
them unchanged. Also the file readers: docx tables and xlsx sheets -> rows of
strings (xlsx un-merges merged cells and formats dates/numbers).
"""
import datetime
import os

import docx
from openpyxl import load_workbook

from parser import ColumnMap, FieldRule, NameSlot, transcribe

# Data lives in a sibling ./data folder, resolved relative to THIS file so the
# repo runs on any machine (no hardcoded absolute paths).
UP = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data") + os.sep
MIX18_DOCX = UP + "MIX18_rooming-list_-_12_06-14_06-_Park_Hotel_Salice_Terme__002__ANON.docx"
PARK_XLSX  = UP + "Copia_di_PARK_HOTEL_SALICE_TERME_040626_agg_rooming_18may_ANON.xlsx"
POLISH_XLSX = UP + "Rooming_List__W\u0142ochy_po_\u0142__30_05__06_06_2026__1__ANON.xlsx"


# ---- file readers ----------------------------------------------------------

def read_docx_rows(path):
    table = docx.Document(path).tables[0]
    return [[c.text.strip() for c in row.cells] for row in table.rows]

def _cell_str(v):
    if v is None:
        return ""
    if isinstance(v, datetime.datetime) or isinstance(v, datetime.date):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

def read_xlsx_rows(path, sheet=0):
    """xlsx -> rows of strings. Un-merges merged ranges (fill-down) and formats dates/ints."""
    ws = load_workbook(path, data_only=True).worksheets[sheet]
    grid = [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]
    for rng in ws.merged_cells.ranges:                       # copy top-left value across merge
        val = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                grid[r - 1][c - 1] = val
    return [[_cell_str(v) for v in row] for row in grid]


# ---- MIX18 (Ukrainian docx): one combined name cell, has passports ---------

def _mix18_role(row):
    return "18" if row[7].strip().upper() == "TOUR-LEADER" else "20"

MIX18_MAP = ColumnMap(
    header_rows=1,
    name_slots=[NameSlot(combined_column=1, name_order="surname_first")],
    role_rule=_mix18_role, default_role="20",
    fields={
        "data_nascita":     FieldRule(column=4, normalizer="dotted_date"),
        "numero_documento": FieldRule(column=5, normalizer="passthrough"),
        "tipo_documento":   FieldRule(column=5, normalizer="doc_type_passport"),
    },
)
def parse_mix18(path=MIX18_DOCX):
    return transcribe(read_docx_rows(path), MIX18_MAP)


# ---- Polish xlsx: separate surname/first-name cols, DOB, blank row 2 -------
# Cols: 0 No., 1 Last Name, 2 First Name, 3 Date of Birth, 4 room.
def _polish_skip(row):
    # Real guests are the numbered rows. This drops the 2-row-tall header (and the
    # phantom copy that merged-cell fill-down creates), blank rows, and the
    # room-type legend in the footer (DBL / TRP / "No. of rooms / people").
    return not (row and row[0].strip().isdigit())

POLISH_MAP = ColumnMap(
    header_rows=1,
    name_slots=[NameSlot(surname_column=1, firstname_column=2)],
    default_role="20",
    skip_row=_polish_skip,
    skip_desc="column_not_digit col0",          # same wording stage 1 compiles, so maps match
    fields={"data_nascita": FieldRule(column=3, normalizer="passthrough")},  # reader already DD/MM/YYYY
)
def parse_polish(path=POLISH_XLSX):
    return transcribe(read_xlsx_rows(path), POLISH_MAP)


# ---- Park Hotel xlsx: TWO people per row, names only, held rows ------------
# Cols: ... 16 Cognome, 17 Nome, 18 Cognome 2, 19 Nome 2 ... (header on row 2).
def _park_skip(row):
    return len(row) > 16 and row[16].strip().startswith("Al.Mat")   # held "names pending" rows

PARK_MAP = ColumnMap(
    header_rows=2,                                    # row 1 summary, row 2 header
    name_slots=[
        NameSlot(surname_column=16, firstname_column=17),   # primary occupant
        NameSlot(surname_column=18, firstname_column=19),   # second occupant (twins) -> 2 guests
    ],
    default_role="20",
    skip_row=_park_skip,
    skip_desc='column_startswith col16 "Al.Mat"',   # same wording stage 1 compiles
    fields={},                                        # no Alloggiati identity fields in this list
)
def parse_park(path=PARK_XLSX):
    return transcribe(read_xlsx_rows(path), PARK_MAP)
